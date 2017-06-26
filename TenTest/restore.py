#! python3
# -*- coding: utf-8 -*-
import re
import os
from openpyxl import load_workbook
"""1. 遍历10份分词语料平均F值最大的，确定学习model
   2. 根据model将大量文本进行机器学习，可以将所有文本加在一起，也可以不加，一般不加
   3. 根据标签逆向恢复斜杠文本"""


def confirm_max_average_f_value():
    max_values = {}
    for root, dirs, files in os.walk(r'D:\TenTest'):
        for directory in dirs:
            if directory.isdigit() and directory != ".idea" and directory != "19113106原始语料":
                directory_path = r'%s\%s' % (root, directory)
                wb = load_workbook(r'%s\%s' % (directory_path, "data.xlsx"))
                ws = wb.active
                print(ws['B13'].value)
                max_values.setdefault(directory, ws['B13'].value)
    print(max_values)
    return max(max_values, key=max_values.get)


def generate_column_structure():
    for root, dirs, files in os.walk(r'D:\TenTest\19113106原始语料'):
        for file in files:
            os.makedirs(r"%s\split" % root, exist_ok=True)
            if file.endswith('.txt'):
                with open(r"%s\%s" % (root, file), 'r', encoding='utf8', errors='ignore') as f:
                    data = f.read()
                    new_string = data.replace('\n', '').replace(' ', '').replace('\t', '').replace('\r', '').replace('。', '。$').replace('；', '；$').replace('！','！$').replace('？', '？$')
                    sentences = new_string.split('$')
                    each_word = []
                    for s in sentences:
                        for word in s:
                            if word == '':  # 跳过文字为空
                                continue
                            elif word == '　':  # 跳过文字为空格
                                continue
                            else:
                                each_word.append(word)
                        each_word.append('')  # 在每个句子后面加个空格
                    print(each_word)
                    all_content = '\n'.join(each_word)
                with open(r"%s\split\split_%s" % (root, file), 'w') as f:
                    f.write(all_content)


def machine_learning():
    for root, dirs, files in os.walk(r'D:\TenTest\19113106原始语料'):
        for file in files:
            if file.startswith('split') and file.endswith('.txt'):
                txt_path = r'%s\%s' % (root, file)  # 学习文本路径
                # print(txt_path)
                # os.makedirs(r"%s\res" % root, exist_ok=True)  # 新建机器学习文本结果目录
                model_path = r'D:\TenTest\0\model'
                output_path = r'%s\res\output_%s' % (root, file)
                print("crf_test -m %s %s > %s" % (model_path, txt_path, output_path))
                command = "crf_test -m %s %s > %s" % (model_path, txt_path, output_path)  # crf命令不能处理带有空格的文件名
                os.system(command)


def reverse_restore():  # 根据标签反向恢复
    for root, dirs, files in os.walk(r'D:\res'):
        for file in files:
            print(file)
            if file.startswith('output_split_') and file.endswith('.txt'):
                print(file)
                # with open(r"%s\%s" % (root, file), 'r', encoding='utf8', errors='ignore') as f:
                #     data = f.read()
                #     print(data)
                # break

if __name__ == '__main__':
    # res = confirm_max_average_f_value()
    # print(res)
    # generate_column_structure()
    # machine_learning()
    reverse_restore()